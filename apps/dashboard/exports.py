"""
Export utilities for Excel and PDF generation
"""
import io
from datetime import datetime, timedelta
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum, Count
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from apps.authentication.decorators import role_required
from apps.payments.models import Caisse, Paiement, SortieCaisse
from apps.orders.models import Commande


@login_required
@role_required(['Radmin', 'Rcomptable'])
def export_ventes_excel(request):
    """Export des ventes en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "La bibliothèque openpyxl n'est pas installée. "
            "Exécutez: pip install openpyxl",
            status=500
        )
    
    # Paramètres de date
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if date_debut:
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
    else:
        date_debut = timezone.now().date() - timedelta(days=30)
    
    if date_fin:
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    else:
        date_fin = timezone.now().date()
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    
    # === Feuille 1: Résumé ===
    ws_resume = wb.active
    ws_resume.title = "Résumé"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-tête
    ws_resume['A1'] = "RAPPORT DES VENTES"
    ws_resume['A1'].font = Font(bold=True, size=16)
    ws_resume['A2'] = f"Période: {date_debut.strftime('%d/%m/%Y')} - {date_fin.strftime('%d/%m/%Y')}"
    ws_resume['A3'] = f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    
    # Statistiques
    paiements = Paiement.objects.filter(
        date_paiement__date__gte=date_debut,
        date_paiement__date__lte=date_fin
    )
    
    depenses = SortieCaisse.objects.filter(
        date_sortie__date__gte=date_debut,
        date_sortie__date__lte=date_fin
    )
    
    total_ventes = paiements.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    nb_commandes = paiements.count()
    
    ws_resume['A5'] = "STATISTIQUES"
    ws_resume['A5'].font = Font(bold=True, size=14)
    
    stats = [
        ("Total des ventes", f"{total_ventes}€"),
        ("Nombre de commandes", nb_commandes),
        ("Panier moyen", f"{(total_ventes / nb_commandes if nb_commandes > 0 else 0):.2f}€"),
        ("Total des dépenses", f"{total_depenses}€"),
        ("Bénéfice net", f"{total_ventes - total_depenses}€"),
    ]
    
    for i, (label, value) in enumerate(stats, start=6):
        ws_resume[f'A{i}'] = label
        ws_resume[f'B{i}'] = value
        ws_resume[f'A{i}'].font = Font(bold=True)
    
    # === Feuille 2: Paiements ===
    ws_paiements = wb.create_sheet("Paiements")
    
    headers = ["Date", "Commande", "Table", "Montant", "Mode de paiement", "Encaissé par"]
    for col, header in enumerate(headers, start=1):
        cell = ws_paiements.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for row, paiement in enumerate(paiements.select_related('commande__table', 'utilisateur'), start=2):
        ws_paiements.cell(row=row, column=1, value=paiement.date_paiement.strftime('%d/%m/%Y %H:%M'))
        ws_paiements.cell(row=row, column=2, value=paiement.commande.numero_commande)
        ws_paiements.cell(row=row, column=3, value=paiement.commande.table.numero_table)
        ws_paiements.cell(row=row, column=4, value=float(paiement.montant))
        ws_paiements.cell(row=row, column=5, value=paiement.get_mode_paiement_display())
        ws_paiements.cell(row=row, column=6, value=paiement.utilisateur.login)
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws_paiements.column_dimensions[get_column_letter(col)].width = 18
    
    # === Feuille 3: Dépenses ===
    ws_depenses = wb.create_sheet("Dépenses")
    
    headers_dep = ["Date", "Type", "Motif", "Montant", "Effectuée par"]
    for col, header in enumerate(headers_dep, start=1):
        cell = ws_depenses.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for row, depense in enumerate(depenses.select_related('type_depense', 'utilisateur'), start=2):
        ws_depenses.cell(row=row, column=1, value=depense.date_sortie.strftime('%d/%m/%Y %H:%M'))
        ws_depenses.cell(row=row, column=2, value=depense.type_depense.nom)
        ws_depenses.cell(row=row, column=3, value=depense.motif)
        ws_depenses.cell(row=row, column=4, value=float(depense.montant))
        ws_depenses.cell(row=row, column=5, value=depense.utilisateur.login)
    
    for col in range(1, len(headers_dep) + 1):
        ws_depenses.column_dimensions[get_column_letter(col)].width = 20
    
    # Générer la réponse
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rapport_ventes_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@role_required(['Radmin', 'Rcomptable'])
def export_ventes_pdf(request):
    """Export des ventes en PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse(
            "La bibliothèque reportlab n'est pas installée. "
            "Exécutez: pip install reportlab",
            status=500
        )
    
    # Paramètres de date
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if date_debut:
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
    else:
        date_debut = timezone.now().date() - timedelta(days=30)
    
    if date_fin:
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    else:
        date_fin = timezone.now().date()
    
    # Données
    paiements = Paiement.objects.filter(
        date_paiement__date__gte=date_debut,
        date_paiement__date__lte=date_fin
    ).select_related('commande__table', 'utilisateur')
    
    depenses = SortieCaisse.objects.filter(
        date_sortie__date__gte=date_debut,
        date_sortie__date__lte=date_fin
    ).select_related('type_depense', 'utilisateur')
    
    total_ventes = paiements.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    nb_commandes = paiements.count()
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        textColor=colors.HexColor('#3B82F6')
    )
    elements.append(Paragraph("RAPPORT DES VENTES", title_style))
    
    # Sous-titre
    subtitle = f"Période: {date_debut.strftime('%d/%m/%Y')} - {date_fin.strftime('%d/%m/%Y')}"
    elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Paragraph(f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Résumé
    elements.append(Paragraph("RÉSUMÉ", styles['Heading2']))
    
    summary_data = [
        ["Indicateur", "Valeur"],
        ["Total des ventes", f"{total_ventes}€"],
        ["Nombre de commandes", str(nb_commandes)],
        ["Panier moyen", f"{(total_ventes / nb_commandes if nb_commandes > 0 else 0):.2f}€"],
        ["Total des dépenses", f"{total_depenses}€"],
        ["Bénéfice net", f"{total_ventes - total_depenses}€"],
    ]
    
    summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E2E8F0')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Liste des paiements
    elements.append(Paragraph("DÉTAIL DES PAIEMENTS", styles['Heading2']))
    
    if paiements.exists():
        payment_data = [["Date", "Commande", "Table", "Montant", "Mode"]]
        for p in paiements[:50]:  # Limiter à 50 pour éviter des PDF trop longs
            payment_data.append([
                p.date_paiement.strftime('%d/%m/%Y'),
                p.commande.numero_commande,
                p.commande.table.numero_table,
                f"{p.montant}€",
                p.get_mode_paiement_display()
            ])
        
        payment_table = Table(payment_data, colWidths=[3*cm, 4*cm, 2.5*cm, 2.5*cm, 3*cm])
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ]))
        elements.append(payment_table)
    else:
        elements.append(Paragraph("Aucun paiement pour cette période.", styles['Normal']))
    
    # Générer le PDF
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"rapport_ventes_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.pdf"
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@role_required(['Radmin', 'Rcomptable'])
def export_commandes_excel(request):
    """Export des commandes en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "La bibliothèque openpyxl n'est pas installée.",
            status=500
        )
    
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    if date_debut:
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
    else:
        date_debut = timezone.now().date() - timedelta(days=30)
    
    if date_fin:
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    else:
        date_fin = timezone.now().date()
    
    commandes = Commande.objects.filter(
        date_commande__date__gte=date_debut,
        date_commande__date__lte=date_fin
    ).select_related('table').order_by('-date_commande')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    headers = ["N° Commande", "Table", "Date", "Montant", "Statut"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row, cmd in enumerate(commandes, start=2):
        ws.cell(row=row, column=1, value=cmd.numero_commande)
        ws.cell(row=row, column=2, value=cmd.table.numero_table)
        ws.cell(row=row, column=3, value=cmd.date_commande.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=4, value=float(cmd.montant_total))
        ws.cell(row=row, column=5, value=cmd.get_statut_display())
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"commandes_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
